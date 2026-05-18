"""
KPI Analyzer — Weekly Report Generator
Connects to PostgreSQL, pulls the week's data, computes KPIs,
and exports a formatted Excel file every Friday at 20:00.

CRON (not yet activated):
    0 20 * * 5  /home/ighor/Coding/.venv/bin/python /home/ighor/Coding/KPIAnalyzer/main.py

Cloud SQL note:
    When migrating to Cloud SQL, just update the .env values.
    If using Cloud SQL Proxy (Unix socket):
        DB_HOST=/cloudsql/project:region:instance
    Or use the Cloud SQL Python connector (cloud-sql-python-connector package).
"""

import os
from datetime import date, timedelta
import pandas as pd
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

load_dotenv()

# ============================================================
# SOURCE TABLE CONFIG — adjust to match your schema
# ============================================================

SOURCE_TABLE = os.getenv("SOURCE_TABLE", "reviews")
COL_ANALYST  = os.getenv("COL_ANALYST",  "analyst_name")
COL_DATE     = os.getenv("COL_DATE",     "reviewed_at")
COL_ERROR    = os.getenv("COL_ERROR",    "is_error")

OUTPUT_DIR   = os.path.join(os.path.dirname(__file__), "exports")


# ============================================================
# CONNECTION (SQLAlchemy — Cloud SQL ready)
# ============================================================

def get_engine():
    host     = os.getenv("DB_HOST", "localhost")
    port     = os.getenv("DB_PORT", "5432")
    dbname   = os.getenv("DB_NAME")
    user     = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")

    url = f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{dbname}"
    return create_engine(url)


# ============================================================
# DATA EXTRACTION — current week (Mon–Fri)
# ============================================================

def week_range() -> tuple[date, date]:
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    friday = monday + timedelta(days=4)
    return monday, friday


def fetch_data(engine) -> pd.DataFrame:
    monday, friday = week_range()

    query = text(f"""
        SELECT
            {COL_ANALYST}  AS analyst_name,
            {COL_DATE}     AS reviewed_at,
            {COL_ERROR}    AS is_error
        FROM {SOURCE_TABLE}
        WHERE {COL_DATE} >= :start
          AND {COL_DATE} <  :end
        ORDER BY {COL_ANALYST}, {COL_DATE}
    """)

    with engine.connect() as conn:
        df = pd.read_sql(query, conn, params={
            "start": monday,
            "end":   friday + timedelta(days=1),
        })

    print(f"  Fetched {len(df)} rows from [{monday}] to [{friday}]")
    return df


# ============================================================
# KPI COMPUTATION
# ============================================================

def compute_kpis(df: pd.DataFrame) -> pd.DataFrame:
    kpi = (
        df.groupby("analyst_name")
        .agg(
            total_reviews=("analyst_name", "count"),
            errors=("is_error", "sum"),
        )
        .reset_index()
    )

    kpi["error_rate"] = (kpi["errors"] / kpi["total_reviews"] * 100).round(2)
    kpi = kpi.rename(columns={
        "analyst_name":  "Analyst Name",
        "total_reviews": "Total Reviews",
        "errors":        "Errors",
        "error_rate":    "Error Rate (%)",
    })

    # Totals row
    totals = {
        "Analyst Name":    "TOTAL",
        "Total Reviews":   kpi["Total Reviews"].sum(),
        "Errors":          kpi["Errors"].sum(),
        "Error Rate (%)":  round(kpi["Errors"].sum() / kpi["Total Reviews"].sum() * 100, 2)
                           if kpi["Total Reviews"].sum() > 0 else 0.0,
    }
    kpi = pd.concat([kpi, pd.DataFrame([totals])], ignore_index=True)
    return kpi


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_excel(kpi: pd.DataFrame, raw: pd.DataFrame):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    monday, friday = week_range()
    filename = f"KPI_Report_{monday}_to_{friday}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # --- Sheet 1: KPI Summary ---
        kpi.to_excel(writer, sheet_name="KPI Summary", index=False)

        ws = writer.sheets["KPI Summary"]
        _format_summary_sheet(ws, kpi)

        # --- Sheet 2: Raw Data ---
        if not raw.empty:
            raw.to_excel(writer, sheet_name="Raw Data", index=False)
            _autofit_columns(writer.sheets["Raw Data"])

    print(f"  Excel saved → {filepath}")
    return filepath


def _format_summary_sheet(ws, kpi: pd.DataFrame):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill  = PatternFill("solid", fgColor="6C3EC0")  # Nubank-ish purple
    totals_fill  = PatternFill("solid", fgColor="D9D9D9")
    red_fill     = PatternFill("solid", fgColor="FFCCCC")
    green_fill   = PatternFill("solid", fgColor="CCFFCC")
    white_font   = Font(bold=True, color="FFFFFF")
    bold         = Font(bold=True)
    center       = Alignment(horizontal="center")
    thin         = Side(style="thin")
    border       = Border(bottom=thin)

    # Header row
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = white_font
        cell.alignment = center

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        is_total = row_idx == len(kpi)
        for cell in row:
            cell.alignment = center
            if is_total:
                cell.fill   = totals_fill
                cell.font   = bold
                cell.border = border

        # Color-code error rate column (col 4)
        rate_cell = row[3]
        if not is_total and rate_cell.value is not None:
            try:
                if float(rate_cell.value) > 5:
                    rate_cell.fill = red_fill
                else:
                    rate_cell.fill = green_fill
            except (TypeError, ValueError):
                pass

    # Column widths
    col_widths = [25, 18, 12, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _autofit_columns(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 55)
    print("   KPI Analyzer — Weekly Report")
    print(f"   Week: {week_range()[0]}  to  {week_range()[1]}")
    print("=" * 55)

    print("\n[1/3] Connecting to database...")
    engine = get_engine()
    print("  Connected.")

    print("\n[2/3] Fetching and computing KPIs...")
    raw = fetch_data(engine)
    kpi = compute_kpis(raw)
    print("\n" + kpi.to_string(index=False))

    print("\n[3/3] Exporting to Excel...")
    export_excel(kpi, raw)

    print("\nDone.")


if __name__ == "__main__":
    main()
